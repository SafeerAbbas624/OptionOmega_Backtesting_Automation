from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementClickInterceptedException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
import pandas as pd
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill
import time
from datetime import datetime as dt, time as dt_time

def setup_driver():
    """Initialize the Chrome driver with appropriate options"""
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 30)
    return driver, wait

def login(driver, wait, username, password):
    """Login to OptionOmega"""
    driver.get("https://optionomega.com/login")
    
    # Wait for and fill in login form
    username_field = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@type="email"]')))
    password_field = driver.find_element(By.XPATH, '//*[@type="password"]')
    
    username_field.send_keys(username)
    password_field.send_keys(password)
    password_field.send_keys(Keys.RETURN)
    
    # Wait for login to complete
    time.sleep(5)

def navigate_to_backtester(driver):
    """Navigate to the backtester page"""
    driver.find_element(By.XPATH, '//button[@class="btn-primary"]').click()
    time.sleep(10)

def input_field(driver, wait, field_id, value):
    """Input a value into a field, with special handling for time inputs"""
    field = wait.until(EC.element_to_be_clickable((By.XPATH, field_id)))
    field.clear()
    
    if field.get_attribute('type') == 'time':
        # Parse the time value
        try:
            parsed_time = dt.strptime(value, "%I:%M %p")
            formatted_time = parsed_time.strftime("%H:%M")
            
            # Input the time
            field.send_keys(formatted_time)
            
            # If it was PM, increment the hours
            ActionChains(driver).send_keys(Keys.ARROW_RIGHT).perform()
            if 'AM' in value:
                for _ in range(3):
                    ActionChains(driver).send_keys(Keys.ARROW_UP).perform()
                    time.sleep(0.1)
            if 'PM' in value:
                for _ in range(2):
                    ActionChains(driver).send_keys(Keys.ARROW_UP).perform()
                    time.sleep(0.1)
            
        except ValueError:
            print(f"Invalid time format: {value}. Expected format: HH:MM AM/PM")
            return
    else:
        # For non-time inputs, just send the value as is
        field.send_keys(str(value))




def click_button(wait, click_id):
    """Click on the button using ActionChains."""

    click_element = wait.until(EC.presence_of_element_located((By.XPATH, click_id)))
    actions = ActionChains(wait._driver)
    actions.move_to_element(click_element).perform()  # Move to element (if necessary)
    actions.click(click_element).perform()




def input_backtest_criteria(driver, wait, row):
    """Input backtest criteria from DataFrame row"""
    # Input basic criteria
    try:
        # input start date
        input_field(driver, wait, '//*[@class="input mt-1"]', row['Start Date'])
        time.sleep(0.5)
    except Exception:
        time.sleep(5)
        try:
            # input start date
            input_field(driver, wait, '//*[@class="input mt-1"]', row['Start Date'])
            time.sleep(0.5)
        except Exception:
            driver.close()
            time.sleep(5)
            login(driver, wait, "Seanseahsg@gmail.com", "$Alpha2024")
            navigate_to_backtester(driver)
            time.sleep(5)
            input_field(driver, wait, '//*[@class="input mt-1"]', row['Start Date'])
            time.sleep(0.5)
    # input end date
    input_field(driver, wait, '//div[@class="relative mt-1"]/input', row['End Date'])
    time.sleep(0.5)
    # dropdown for ticker
    click_button(wait, '//*[@id="headlessui-listbox-button-12"]')
    time.sleep(0.5)
    if 'SPX' in row['Ticker']:
        time.sleep(0.5)
        click_button(wait, '/html/body/div[2]/div/div/div/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[1]/div/div/ul/li[1]')
    elif 'SPY' in row['Ticker']:
        time.sleep(0.5)
        click_button(wait, '/html/body/div[2]/div/div/div/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[1]/div/div/ul/li[2]')
    elif 'IWM' in row['Ticker']:
        time.sleep(0.5)
        click_button(wait, '/html/body/div[2]/div/div/div/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[1]/div/div/ul/li[3]')
    elif 'QQQ' in row['Ticker']:
        time.sleep(0.5)
        click_button(wait, '/html/body/div[2]/div/div/div/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[1]/div/div/ul/li[4]')
    elif 'AAPL' in row['Ticker']:
        time.sleep(0.5)
        click_button(wait, '/html/body/div[2]/div/div/div/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[1]/div/div/ul/li[5]')
    elif 'TSLA' in row['Ticker']:
        time.sleep(0.5)
        click_button(wait, '/html/body/div[2]/div/div/div/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[1]/div/div/ul/li[6]')
        time.sleep(0.5)
    
    
    # adding leg options
    # Input leg details
    # first leg
    click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/div[2]/button')
    time.sleep(0.5)
    # Select buy/sell
    if 'Buy' in row['Sell/ Buy1']:
        click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr/td/div[1]/button[2]')
        time.sleep(0.5)
    # Select call/put
    if 'Call' in row['Call/ Put1']:
        click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr/td/div[1]/button[3]')
        time.sleep(0.5)
    # Input quantity
    input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr/td/div[2]/div/input', row['QTY1'])
    time.sleep(0.5)
    # Input delta
    input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr/td/div[3]/div/input', row['Delta1'])
    time.sleep(0.5)
    # Input DTE
    input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr/td/div[4]/div/input', row['DTE1'])
    time.sleep(0.5)



    if row['Sell/ Buy2'] and row['Call/ Put2']:
        # second leg
        if 'Yes' in row['Attach to Leg 1']:
            # check for attach to leg and click
            click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr/td/div[5]/button[1]')
            time.sleep(0.5)
            # Second leg 
            # Select buy/sell
            if 'Buy' in row['Sell/ Buy2']:
                click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[1]/button[2]')
                time.sleep(0.5)
                # Select call/put
            if 'Call' in row['Call/ Put2']:
                click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[1]/button[3]')
                time.sleep(0.5)
            # Input quantity
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[2]/div/input', row['QTY2'])
            time.sleep(0.5)
            # Input delta
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[3]/div/input', row['Delta2'])
            time.sleep(0.5)
            # Input DTE
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[4]/div/input', row['DTE2'])
            time.sleep(0.5)
        else:
                # click on add leg button
            click_button(wait, '//div[@class="mt-2 ml-1 link w-full"]/button')
            time.sleep(0.5)
            # Second leg 
            # Select buy/sell
            if 'Buy' in row['Sell/ Buy2']:
                click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[1]/button[2]')
                time.sleep(0.5)
                # Select call/put
            if 'Call' in row['Call/ Put2']:
                click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[1]/button[3]')
                time.sleep(0.5)
            # Input quantity
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[2]/div/input', row['QTY2'])
            time.sleep(0.5)
            # Input delta
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[3]/div/input', row['Delta2'])
            time.sleep(0.5)
            # Input DTE
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[2]/td/div[4]/div/input', row['DTE2'])
            time.sleep(0.5)
    else:
        pass



    if row['Sell/ Buy3'] and row['Call/ Put3']:
        # third leg
        # click on add leg button
        click_button(wait, '//div[@class="mt-2 ml-1 link w-full"]/button')
        time.sleep(0.5)
        # Select buy/sell
        if 'Buy' in row['Sell/ Buy3']:
            click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[3]/td/div[1]/button[2]')
            time.sleep(0.5)
            # Select call/put
        if 'Call' in row['Call/ Put3']:
            click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[3]/td/div[1]/button[3]')
            time.sleep(0.5)
        # Input quantity
        input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[3]/td/div[2]/div/input', row['QTY3'])
        time.sleep(0.5)
        # Input delta
        input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[3]/td/div[3]/div/input', row['Delta3'])
        time.sleep(0.5)
        # Input DTE
        input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[3]/td/div[4]/div/input', row['DTE3'])
        time.sleep(0.5)
    else:
        pass



    if row['Sell/ Buy4'] and row['Call/ Put4']:
        # forth leg 
        if 'Yes' in row['Attach to Leg 3']:
            # check for attach to leg and click
            click_button(wait, '(//button[@class="bg-transparent ml-3 pl-3 pr-3 py-1 focus:outline-none focus:ring-transparent border border-gray-500 rounded-md"])[4]')
            time.sleep(0.5)
            # 4th leg 
            # Select buy/sell
            if 'Buy' in row['Sell/ Buy4']:
                click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[1]/button[2]')
                time.sleep(0.5)
            # Select call/put
            if 'Call' in row['Call/ Put4']:
                click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[1]/button[3]')
                time.sleep(0.5)
            # Input quantity
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[2]/div/input', row['QTY4'])
            time.sleep(0.5)
            # Input delta
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[3]/div/input', row['Delta4'])
            time.sleep(0.5)
            # Input DTE
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[4]/div/input', row['DTE4'])
            time.sleep(0.5)
        else:
            # click on add leg button
            click_button(wait, '//div[@class="mt-2 ml-1 link w-full"]/button')
            time.sleep(0.5)
            # 4th leg 
            # Select buy/sell
            if 'Buy' in row['Sell/ Buy4']:
                click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[1]/button[2]')
                time.sleep(0.5)
            # Select call/put
            if 'Call' in row['Call/ Put4']:
                click_button(wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[1]/button[3]')
                time.sleep(0.5)
            # Input quantity
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[2]/div/input', row['QTY4'])
            time.sleep(0.5)
            # Input delta
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[3]/div/input', row['Delta4'])
            time.sleep(0.5)
            # Input DTE
            input_field(driver, wait, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[1]/div[2]/div/div[2]/div[2]/div[4]/table/tbody/tr[4]/td/div[4]/div/input', row['DTE4'])
            time.sleep(0.5)
    else:
        pass

    
    
    # Input trade settings
    # Starting funds
    input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[1]', row['Starting Funds'])
    time.sleep(0.5)
    # Margin allocation %
    input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[2]', row['Margin Allocation % Per Trade'])
    time.sleep(0.5)
    # max open trade
    input_field(driver, wait, '(//div[@class="mt-1"]/input)[1]', row['Max Open Trades'])
    time.sleep(0.5)
    # Max Contracts Per Trade
    input_field(driver, wait, '(//div[@class="mt-1"]/input)[2]', row['Max Contracts Per Trade'])
    time.sleep(0.5)
    # Ignore Margin Requirement
    if 'Yes' in row['Ignore Margin Requirement']:
        click_button(wait, '//button[@id="headlessui-switch-21"]')
        time.sleep(0.5)
    else:
        pass
    # Max Allocation Amount Per Trade
    input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[3]', row['Max Allocation Amount Per Trade'])
    time.sleep(0.5)

    # Entry Conditions
    # Entry Time
    
    if  not row['Entry Time']:
        # Floting entry time
        click_button(wait, '//*[@id="headlessui-switch-23"]')
        time.sleep(0.5)
        # minimum entry time
        input_field(driver, wait, '(//div[@class="mt-1"]/input)[3]', row['Min Entry Time'])
        time.sleep(0.5)
        # maximum entry time
        input_field(driver, wait, '(//div[@class="mt-1"]/input)[4]', row['Max Entry Time'])
        time.sleep(0.5)
    else:
        # entry time if floating is no
        input_field(driver, wait, '(//div[@class="mt-1"]/input)[3]', str(row['Entry Time']))
        time.sleep(0.3)

    # dropdown for frequency
    # weekly entries
    click_button(wait, '//*[@id="headlessui-listbox-button-25"]')
    time.sleep(0.5)
    if 'Daily' in row['Frequency']:
        click_button(wait, '//span[contains(text(),"Daily")]')

    elif 'Weekly' in row['Frequency']:
        click_button(wait, '//span[contains(text(),"Weekly")]')

    elif 'Specific Dates' in row['Frequency']:
        click_button(wait, '//span[contains(text(),"Specific Dates")]')
    
    if 'Weekly' in row['Frequency']:
        if 'M' in row['Days If Weekly']:
            click_button(wait, '//button[contains(text(),"M")]')
            time.sleep(0.5)

        if 'Tu' in row['Days If Weekly']:
            click_button(wait, '//button[contains(text(),"Tu")]')
            time.sleep(0.5)

        if 'W' in row['Days If Weekly']:
            click_button(wait, '//button[contains(text(),"W")]')
            time.sleep(0.5)

        if 'Th' in row['Days If Weekly']:
            click_button(wait, '//button[contains(text(),"Th")]')
            time.sleep(0.5)

        if 'F' in row['Days If Weekly']:
            click_button(wait, '//button[contains(text(),"F")]')
            time.sleep(0.5)

    # specific dates
    if 'Specific Dates' in row['Frequency']:
        click_button(wait, '//textarea[@class="input mt-1"]')
        time.sleep(0.5)
        input_field(driver, wait, '//textarea[@class="input mt-1"]', row['Dates if Specific Dates'])
        time.sleep(0.5)

    if 'Yes' in row['Use VIX']:
        # click on use VIX
        click_button(wait, '//button[@id="headlessui-switch-27"]')
        time.sleep(0.5)
        if 'Yes' in row['Use Floating Entry Time']:
            input_field(driver, wait, '(//div[@class="mt-1"]/input)[5]', row['Min VIX'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1"]/input)[6]', row['MAX VIX'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[4]', row['Min VIX Overnight Move Up'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[5]', row['Max VIX Overnight Move Up'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[6]', row['Min VIX Overnight Move Down'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[7]', row['Max VIX Overnight Move Down'])
            time.sleep(0.5)
        else:
            input_field(driver, wait, '(//div[@class="mt-1"]/input)[4]', row['Min VIX'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1"]/input)[5]', row['MAX VIX'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[4]', row['Min VIX Overnight Move Up'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[5]', row['Max VIX Overnight Move Up'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[6]', row['Min VIX Overnight Move Down'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[7]', row['Max VIX Overnight Move Down'])
            time.sleep(0.5)
    else:
        click_button(wait, '//button[@id="headlessui-switch-27"]')
        time.sleep(0.5)

    if 'Yes' in row['Use Technical Indicators']:
        click_button(wait, '//button[@id="headlessui-switch-29"]')
        time.sleep(0.5)
        if 'Yes' in row['Use Floating Entry Time']:
            input_field(driver, wait, '(//div[@class="mt-1"]/input)[9]', row['Min RSI'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1"]/input)[10]', row['Max RSI'])
            time.sleep(0.5)
            if row['SMA Entry']:
                if 'None' in row['SMA Entry']:
                    pass
                else:
                    click_button(wait, '(//div[@class="mt-1 relative"]/button)[3]')
                    if 'Below SMA' in row['SMA Entry']:
                        click_button(wait, '//span[contains(text(),"Below SMA")]')
                    elif 'Above SMA' in row['SMA Entry']:
                        click_button(wait, '//span[contains(text(),"Above SMA")]')
                    elif 'Compare SMA' in row['SMA Entry']:
                        click_button(wait, '//span[contains(text(),"Compare SMA")]')
                    time.sleep(0.5)
                    if row['Above/ Below SMA DAYS']:
                        click_button(wait, '(//div[@class="mt-1 relative"]/button)[4]')
                        click_button(wait, f'//span[contains(text(),"{row["Above/ Below SMA DAYS"]}")]')

            if row['EMA Entry']:
                click_button(wait, '(//div[@class="mt-1 relative"]/button)[5]')
                time.sleep(0.5)
                if 'Below EMA' in row['EMA Entry']:
                    click_button(wait, '//span[contains(text(),"Below EMA")]')
                    time.sleep(0.5)
                elif 'Above EMA' in row['EMA Entry']:
                    click_button(wait, '//span[contains(text(),"Above EMA")]')
                    time.sleep(0.5)
                elif 'Compare EMA' in row['EMA Entry']:
                    click_button(wait, '//span[contains(text(),"Compare EMA")]')
                    time.sleep(0.5)

                if row['Minutes']:
                    click_button(wait, '(//div[@class="mt-1 relative"]/button)[6]')
                    time.sleep(0.5)
                    click_button(wait, f'//span[contains(text(),"{row["Minutes"]}")]')


        else:
            input_field(driver, wait, '(//div[@class="mt-1"]/input)[8]', row['Min RSI'])
            time.sleep(0.5)
            input_field(driver, wait, '(//div[@class="mt-1"]/input)[9]', row['Max RSI'])
            time.sleep(0.5)
            if row['SMA Entry']:
                if 'None' in row['SMA Entry']:
                    pass
                else:
                    click_button(wait, '(//div[@class="mt-1 relative"]/button)[3]')
                    if 'Below SMA' in row['SMA Entry']:
                        click_button(wait, '//span[contains(text(),"Below SMA")]')
                    elif 'Above SMA' in row['SMA Entry']:
                        click_button(wait, '//span[contains(text(),"Above SMA")]')
                    elif 'Compare SMA' in row['SMA Entry']:
                        click_button(wait, '//span[contains(text(),"Compare SMA")]')
                    time.sleep(0.5)
                    if row['Above/ Below SMA DAYS']:
                        click_button(wait, '(//div[@class="mt-1 relative"]/button)[4]')
                        click_button(wait, f'//span[contains(text(),"{row["Above/ Below SMA DAYS"]}")]')

            if row['EMA Entry']:
                click_button(wait, '(//div[@class="mt-1 relative"]/button)[5]')
                time.sleep(0.5)
                if 'Below EMA' in row['EMA Entry']:
                    click_button(wait, '//span[contains(text(),"Below EMA")]')
                    time.sleep(0.5)
                elif 'Above EMA' in row['EMA Entry']:
                    click_button(wait, '//span[contains(text(),"Above EMA")]')
                    time.sleep(0.5)
                elif 'Compare EMA' in row['EMA Entry']:
                    click_button(wait, '//span[contains(text(),"Compare EMA")]')
                    time.sleep(0.5)

                if row['Minutes']:
                    click_button(wait, '(//div[@class="mt-1 relative"]/button)[6]')
                    time.sleep(0.5)
                    click_button(wait, f'//span[contains(text(),"{row["Minutes"]}")]')
    else:
        click_button(wait, '//button[@id="headlessui-switch-29"]')
        time.sleep(0.5)

    if 'yes' in row['Use Gaps']:
        click_button(wait, '//button[@id="headlessui-switch-31"]')
        time.sleep(0.5)
        input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[8]', row['Min Gap Up'])
        time.sleep(0.5)
        input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[9]', row['Max Gap Up'])
        time.sleep(0.5)
        input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[10]', row['Min Gap Down'])
        time.sleep(0.5)
        input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[11]', row['Max Gap Down'])
        time.sleep(0.5)
    else:
        click_button(wait, '//button[@id="headlessui-switch-31"]')
        time.sleep(0.5)

    # use leg group
    if 'Yes' in row['Use Leg Groups']:
        click_button(wait, '//*[@id="headlessui-switch-403"]')
        time.sleep(0.5)
    

    input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[12]', row['Profit Target (%)'])
    time.sleep(0.5)

    input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[13]', row['Stop Loss (%)'])
    time.sleep(0.5)

    # Exit Conditions
    if 'Yes' in row['Use Early Exit']:
        click_button(wait, '//*[@id="headlessui-switch-45"]')
        time.sleep(0.5)
        input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[14]', row['DTE'])
        time.sleep(0.5)
        input_field(driver, wait, '(//input[@class="input mt-1"])[2]', row['Early Exit Time'])
        time.sleep(0.5)
    else:
        click_button(wait, '//*[@id="headlessui-switch-45"]')
        time.sleep(0.5)

    # use MISC.
    if 'Yes' in row['Use Commissions & Fees']:
        click_button(wait, '//*[@id="headlessui-switch-55"]')
        time.sleep(0.5)
        input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[15]', row['Per Contract Opening Commissions & Fees'])
        time.sleep(0.5)
        input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[16]', row['Per Contract Closing Commissions & Fees'])
        time.sleep(0.5)


    if 'Yes' in row['Use Slippage']:
        click_button(wait, '//*[@id="headlessui-switch-57"]')
        time.sleep(0.5)
        input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[17]', row['Entry Slippage'])
        time.sleep(0.5)
        input_field(driver, wait, '(//div[@class="mt-1 relative"]/input)[18]', row['Exit Slippage'])
        time.sleep(0.5)


    if 'Yes' in row['Ignore Trades with Wide Bid-Ask Spread']:
        click_button(wait, '//*[@id="headlessui-switch-59"]')
        time.sleep(0.5)
    
    if 'Yes' in row['Use Blackout Days']:
        click_button(wait, '//*[@id="headlessui-switch-65"]')
        time.sleep(0.5)
        if 'Specific Dates' in row['Frequency']:
            input_field(driver, wait, '(//textarea[@class="input mt-1"])[2]', row['Blackout Days'])
            time.sleep(0.5)
        else:
            input_field(driver, wait, '//textarea[@class="input mt-1"]', row['Blackout Days'])
            time.sleep(0.5)
    
    # Additional settings based on your Excel headers
    # Add more input fields as needed



def run_backtest(wait):
    """Run the backtest and wait for results"""
    run_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="headlessui-dialog-8"]/div/div[2]/div/form/div[2]/button[2]'))) #  //button[@class="ml-4 btn-primary"]
    run_button.click()
    
    # Wait for results to load
    time.sleep(30)

def get_element_text(wait, element_id):
    """Get text from an element"""
    element = wait.until(EC.element_to_be_clickable((By.XPATH, element_id)))
    return element.text

def extract_results(driver, wait, row):
    """Extract backtest results"""
    results = {}
    
    # Extract basic metrics with individual try-except blocks
    try:
        results['P/L'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[1]')
    except Exception:
        results['P/L'] = ''
    
    try:
        results['CAGR'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[2]')
    except Exception:
        results['CAGR'] = ''
    
    try:
        results['Max Drawdown'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[3]')
    except Exception:
        results['Max Drawdown'] = ''
    
    try:
        results['MAR Ratio'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[4]')
    except Exception:
        results['MAR Ratio'] = ''
    
    try:
        results['Win Percentage'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[5]')
    except Exception:
        results['Win Percentage'] = ''
    
    try:
        results['Total Premium'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[6]')
    except Exception:
        results['Total Premium'] = ''
    
    try:
        capture_rate = get_element_text(wait, '(//dt[@class="text-sm font-medium text-white truncate"])[12]')
        if 'Capture Rate' in capture_rate:
            results['Capture Rate'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[7]')
        else:
            results['Capture Rate'] = 0
    except Exception:
        results['Capture Rate'] = ''
    
    try:
        starting_capital = get_element_text(wait, '(//dt[@class="text-sm font-medium text-white truncate"])[13]')
        if 'Starting Capital' in starting_capital:
            results['Starting Capital'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[7]')
        else:
            results['Starting Capital'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[8]')
    except Exception:
        results['Starting Capital'] = ''
    
    try:
        ending_capital = get_element_text(wait, '(//dt[@class="text-sm font-medium text-white truncate"])[14]')
        if 'Ending Capital' in ending_capital:
            results['Ending Capital'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[8]')
        else:
            results['Ending Capital'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[9]')
    except Exception:
        results['Ending Capital'] = ''
    
    try:
        results['Avg Per Trade'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[10]')
    except Exception:
        results['Avg Per Trade'] = ''
    
    try:
        results['Avg Winner'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[11]')
    except Exception:
        results['Avg Winner'] = ''
    
    try:
        results['Avg Loser'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[12]')
    except Exception:
        results['Avg Loser'] = ''
    
    try:
        results['Max Winner'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[13]')
    except Exception:
        results['Max Winner'] = ''
    
    try:
        results['Max Loser'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[14]')
    except Exception:
        results['Max Loser'] = ''
    
    try:
        results['Avg Days in Trade'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[15]')
    except Exception:
        results['Avg Days in Trade'] = ''
    
    try:
        results['Trades'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[16]')
    except Exception:
        results['Trades'] = ''
    
    try:
        results['Winners'] = get_element_text(wait, '(//dd[@class="mt-1 text-lg font-semibold text-white"])[17]')
    except Exception:
        results['Winners'] = ''

    try:
        click_button(wait, '(//div[@class="has-tooltip inline-flex"])[2]')
        time.sleep(8)
        input_field(driver, wait, '//div[@class="mt-1"]/input', row['Name'])
        time.sleep(0.5)
        click_button(wait, '//div[@class="flex items-center py-1 pb-6"]/button')
        time.sleep(0.5)
        click_button(wait, '//*[@class="ml-4 btn-primary" and @type="submit"]')
        time.sleep(7)
        try:
            link_element = driver.find_element(By.XPATH, '//span[@class="has-tooltip"]/a')
            results['Links'] = link_element.get_attribute('href')
        except Exception:
            results['Links'] = "Link Not Found On Page"
    except Exception:
        time.sleep(5)
        try:
            click_button(wait, '(//div[@class="has-tooltip inline-flex"])[2]')
            time.sleep(8)
            input_field(driver, wait, '//div[@class="mt-1"]/input', row['Name'])
            time.sleep(0.5)
            click_button(wait, '//div[@class="flex items-center py-1 pb-6"]/button')
            time.sleep(0.5)
            click_button(wait, '//*[@class="ml-4 btn-primary" and @type="submit"]')
            time.sleep(7)
            try:
                link_element = driver.find_element(By.XPATH, '//span[@class="has-tooltip"]/a')
                results['Links'] = link_element.get_attribute('href')
            except Exception:
                results['Links'] = "Link Not Found On Page"
        except Exception:
            time.sleep(2)
            click_button(wait, '//*[@id="headlessui-menu-button-1"]')
            time.sleep(1)
            click_button(wait, '(//a[@class="block px-4 py-2 text-sm text-white"])[2]')
            time.sleep(5)
            login(driver, wait, "Seanseahsg@gmail.com", "$Alpha2024")
            return results

    time.sleep(2)
    click_button(wait, '//*[@id="headlessui-menu-button-1"]')
    time.sleep(1)
    click_button(wait, '(//a[@class="block px-4 py-2 text-sm text-white"])[2]')
    time.sleep(5)
    login(driver, wait, "Seanseahsg@gmail.com", "$Alpha2024")
    
    return results




def process_input_file(driver, wait, input_file):
    """Process the input Excel file and write results to a separate output file"""
    # Read the Excel file, including the first two rows
    df = pd.read_excel(input_file, header=None, keep_default_na=False, na_values=[])
    
    # Extract main header and sub-header
    main_header = df.iloc[0].tolist()
    sub_header = df.iloc[1].tolist()
    
    # Read the data, using the second row as headers
    data_df = pd.read_excel(input_file, header=1, keep_default_na=False, na_values=[])
    
    # Function to convert float to int if possible
    def float_to_int(x):
        return int(x) if isinstance(x, float) and x.is_integer() else x

    # Apply the conversion to all columns
    data_df = data_df.applymap(float_to_int)
    
    # Convert date columns to datetime without time
    date_columns = ['Start Date', 'End Date']  # Add other date columns if needed
    for col in date_columns:
        if col in data_df.columns:
            data_df[col] = pd.to_datetime(data_df[col]).dt.date
    
    # Identify time columns (adjust this list based on your actual column names)
    time_columns = ['Entry Time', 'Min Entry Time', 'Max Entry Time', 'Early Exit Time']
    
    # Function to parse and format time strings
    def parse_and_format_time(time_str):
        if pd.isna(time_str):
            return None
        if isinstance(time_str, dt_time):
            return time_str.strftime("%I:%M %p")
        try:
            # Try parsing as "HH:MM:SS" or "HH:MM"
            parsed_time = dt.strptime(time_str, "%H:%M:%S" if ":" in time_str else "%H:%M")
            return parsed_time.strftime("%I:%M %p")
        except ValueError:
            try:
                # Try parsing as "HH:MM AM/PM"
                return dt.strptime(time_str, "%I:%M %p").strftime("%I:%M %p")
            except ValueError:
                print(f"Unable to parse time: {time_str}")
                return time_str  # Return original string if unable to parse

    # Convert time columns to formatted time strings
    for col in time_columns:
        if col in data_df.columns:
            data_df[col] = data_df[col].apply(parse_and_format_time)
    
    # Create output file name
    output_file = f"output_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Add "Results" to the main header
    main_header.extend(["Results"])
    
    # Create a placeholder for the results sub-header
    results_sub_header = [""] * (len(sub_header) - len(main_header) + 1) + ['P/L']+ ['CAGR']+ ['Max Drawdown']+ ['MAR Ratio']+ ['Win Percentage']+ ['Total Premium']+ ['Capture Rate']+ ['Starting Capital']+ ['Ending Capital']+ ['Avg Per Trade']+ ['Avg Winner']+ ['Avg Loser']+ ['Max Winner']+ ['Max Loser']+ ['Avg Days in Trade']+ ['Trades']+ ['Winners']+['Sharable Links']
    sub_header.extend(results_sub_header)
    
    # Write the main header and sub-header to the output file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        pd.DataFrame([main_header, sub_header]).to_excel(writer, index=False, header=False, startrow=0)
    
    results_keys = []  # To store the keys of the results dictionary
    
    for index, row in data_df.iterrows():
        print(f'Index: {index} \nRow : {row}')
        navigate_to_backtester(driver)
        input_backtest_criteria(driver, wait, row)
        run_backtest(wait)
        results = extract_results(driver, wait, row)
        
        # Store the keys of the results dictionary
        if not results_keys:
            results_keys = list(results.keys())
        
        # Combine input data with results
        combined_row = pd.concat([row, pd.Series(results)], axis=0)
        
        # Append the combined row to the output file
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            combined_row.to_frame().T.to_excel(writer, index=False, header=False, startrow=index+2)
    
    # Update the results sub-header with the actual keys
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    for col, key in enumerate(results_keys, start=len(sub_header) - len(results_keys) + 1):
        ws.cell(row=2, column=col, value=key)
    
    # Adjust column widths and number formats
    time_style = NamedStyle(name='time_style')
    time_style.number_format = 'hh:mm AM/PM'
    
    # Define fill colors
    light_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row_num = row[0].row
        fill = light_green if row_num % 2 == 1 else white
        for cell in row:
            cell.fill = fill
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set number format for numeric columns
        if isinstance(column[2].value, (int, float)):
            for cell in column[2:]:
                cell.number_format = '#,##0.00' if isinstance(cell.value, float) else '#,##0'
        
        # Set date format for date columns
        if column[1].value in date_columns:
            for cell in column[2:]:
                cell.number_format = 'YYYY-MM-DD'
        
        # Set time format for time columns
        if column[1].value in time_columns:
            for cell in column[2:]:
                cell.style = time_style
    
    wb.save(output_file)
    return output_file


def main():
    # Get input file name from user
    input_file = 'input.xlsx'  # Replace with your input file name
    
    # Initialize automation
    driver, wait = setup_driver()
    # try:
    login(driver, wait, "Seanseahsg@gmail.com", "$Alpha2024")
    output_file = process_input_file(driver, wait, input_file)
    print(f"Results saved to: {output_file}")
    # except Exception as e:
    #     print(f"An error occurred: {str(e)}")
    # finally:
    if driver:
        driver.quit()

if __name__ == "__main__":
    main()